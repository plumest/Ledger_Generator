import openpyxl
from formula import *
from randdataS import *
from random import randint

def write(sheet, month, gender):
    if month == 'sep':
        stop = 35
    else:
        stop = 36
    
    for i in range(5, stop):
        food(sheet, i)
        travel(sheet, i)
        if randint(0, 15) == 0:
            education(sheet, i)
        if randint(0, 25) == 0:
            personal(sheet, i, gender)
        if randint(0, 25) == 0:
            enterainment(sheet, i)
        if randint(0, 25) == 0:
            other(sheet, i)


# Formula
def summary(sheet):
    expense(sheet)
    remaining(sheet)
    total(sheet)
    oct_summary(sheet)

customer_name = input("Enter your name (without any special character or space): ")
file_path = f"customer/{customer_name}.xlsx"

# gender
gender = ''
while (gender not in ['m', 'f']):
    gender = input('Enter your gender (M)ale/(F)emale: ').lower()

print('~~ Fill an interger!!!')
salary = int(input('Enter your salary: '))
rent = int(input('Enter your Rent cost: '))
net = int(input('Enter your internet cost: '))
utility = int(input('Enter your utility cost (water/electricity bills): '))

# Open file change all fomular to value
wb = openpyxl.load_workbook('prototype.xlsx')

# ws for September sheet
wb.active = 1
ws = wb.active

# Mock Sep
ws['C5'] = salary
write(ws, 'sep', gender)
monthly(ws, rent, net, utility)
summary(ws)
sep_summary(ws)

# Save file
wb.save(file_path)

# wo for October sheet
wbb = openpyxl.load_workbook(file_path)
wbb.active = 2
wo = wbb.active

# Mock Oct
wo['C5'] = salary
write(wo, 'oct', gender)
monthly(wo, rent, net, utility)
summary(wo)

# Save file
wbb.save(file_path)